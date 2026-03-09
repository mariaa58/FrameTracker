import sys
import os
import json
import cv2
import qdarktheme
from openpyxl import Workbook
from datetime import datetime

from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QImage, QPixmap, QKeySequence
from PyQt5.QtWidgets import (
    QApplication, QWidget, QLabel, QPushButton, QVBoxLayout, QHBoxLayout,
    QFileDialog, QMessageBox, QInputDialog, QTableWidget, QTableWidgetItem,
    QHeaderView, QShortcut
)


class App(QWidget):
    def __init__(self):
        super().__init__()

        self.cap = None
        self.path = None
        self.idx = 0
        self.total = 0
        self.fps = 0
        self.session_dir = None
        self.modificado = False
        self.caminho_projeto = None

        self.setWindowTitle("Visualizador de Frames")
        self.resize(1400, 750)

        self.lblVideo = QLabel("Nenhum vídeo carregado")
        self.lblVideo.setAlignment(Qt.AlignCenter)
        self.lblVideo.setFixedSize(1040, 560)
        self.lblVideo.setStyleSheet("background:black; color:white; font-size:14px;")

        self.lblInfo = QLabel("Frame: 0 / 0     Tempo: 0.0000 s")
        self.lblInfo.setAlignment(Qt.AlignCenter)
        self.lblInfo.setStyleSheet("font-size:14px;")

        self.btnAbrir = QPushButton("Abrir vídeo")
        self.btnAnt = QPushButton("Anterior")
        self.btnProx = QPushButton("Próximo")
        self.btnIr = QPushButton("Ir para...")
        self.btnIni = QPushButton("Marcar início")
        self.btnFim = QPushButton("Marcar fim")
        self.btnCom = QPushButton("Comentário")
        self.btnFrames = QPushButton("Salvar frames")
        self.btnXlsx = QPushButton("Salvar anotações")
        self.btnProjSalvar = QPushButton("Salvar projeto")
        self.btnProjAbrir = QPushButton("Abrir projeto")

        botoes = [
            self.btnAbrir, self.btnAnt, self.btnProx, self.btnIr,
            self.btnIni, self.btnFim, self.btnCom,
            self.btnFrames, self.btnXlsx,
            self.btnProjSalvar, self.btnProjAbrir
        ]
        for botao in botoes:
            botao.setStyleSheet("font-size:14px;")

        self.botoes_video = [
            self.btnAnt, self.btnProx, self.btnIr,
            self.btnIni, self.btnFim, self.btnCom,
            self.btnFrames, self.btnXlsx, self.btnProjSalvar
        ]
        self.habilitar_botoes_video(False)

        self.btnAbrir.clicked.connect(self.abrir)
        self.btnAnt.clicked.connect(self.ant)
        self.btnProx.clicked.connect(self.prox)
        self.btnIr.clicked.connect(self.ir)
        self.btnIni.clicked.connect(self.marcar_ini)
        self.btnFim.clicked.connect(self.marcar_fim)
        self.btnCom.clicked.connect(self.marcar_com)
        self.btnFrames.clicked.connect(self.salvar_frames)
        self.btnXlsx.clicked.connect(self.salvar_xlsx)
        self.btnProjSalvar.clicked.connect(self.salvar_projeto)
        self.btnProjAbrir.clicked.connect(self.abrir_projeto)

        self.tab = QTableWidget(0, 4)
        self.tab.setHorizontalHeaderLabels(["Tipo", "Frame", "Tempo (s)", "Comentário"])

        cabecalho = self.tab.horizontalHeader()
        cabecalho.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        cabecalho.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        cabecalho.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        cabecalho.setSectionResizeMode(3, QHeaderView.Stretch)

        self.tab.setToolTip("Selecione uma anotação e pressione Delete para remover.")

        atalho_delete = QShortcut(QKeySequence(Qt.Key_Delete), self.tab)
        atalho_delete.activated.connect(self.apagar_linha)

        linha_navegacao = QHBoxLayout()
        linha_navegacao.addWidget(self.btnAnt)
        linha_navegacao.addWidget(self.btnProx)
        linha_navegacao.addWidget(self.btnIr)

        coluna_esquerda = QVBoxLayout()
        coluna_esquerda.addWidget(self.lblVideo)
        coluna_esquerda.addWidget(self.lblInfo)
        coluna_esquerda.addLayout(linha_navegacao)

        coluna_botoes = QVBoxLayout()
        coluna_botoes.addWidget(self.btnAbrir)
        coluna_botoes.addSpacing(10)
        coluna_botoes.addWidget(self.btnIni)
        coluna_botoes.addWidget(self.btnFim)
        coluna_botoes.addWidget(self.btnCom)
        coluna_botoes.addSpacing(10)
        coluna_botoes.addWidget(self.btnFrames)
        coluna_botoes.addWidget(self.btnXlsx)
        coluna_botoes.addSpacing(10)
        coluna_botoes.addWidget(self.btnProjSalvar)
        coluna_botoes.addWidget(self.btnProjAbrir)
        coluna_botoes.addStretch()

        coluna_direita = QVBoxLayout()
        coluna_direita.addLayout(coluna_botoes)
        coluna_direita.addWidget(self.tab)

        layout_principal = QHBoxLayout()
        layout_principal.addLayout(coluna_esquerda, 5)
        layout_principal.addLayout(coluna_direita, 2)
        self.setLayout(layout_principal)

        QShortcut(QKeySequence(Qt.Key_Left), self, activated=self.ant)
        QShortcut(QKeySequence(Qt.Key_Right), self, activated=self.prox)
        QShortcut(QKeySequence("I"), self, activated=self.marcar_ini)
        QShortcut(QKeySequence("F"), self, activated=self.marcar_fim)
        QShortcut(QKeySequence("C"), self, activated=self.marcar_com)

        self.autosave_timer = QTimer(self)
        self.autosave_timer.timeout.connect(self._salvar_autosave)
        self.autosave_timer.start(3 * 60 * 1000)

    def closeEvent(self, event):
        if getattr(self, "modificado", False):
            resposta = QMessageBox.question(
                self,
                "Sair",
                "Existem anotações/projeto não salvos.\nDeseja realmente sair?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if resposta == QMessageBox.No:
                event.ignore()
                return

        if self.cap:
            self.cap.release()
            self.cap = None

        event.accept()

    def habilitar_botoes_video(self, ativar=True):
        for botao in self.botoes_video:
            botao.setEnabled(ativar)

    def _criar_pasta_sessao(self):
        if not self.path:
            return

        if self.session_dir and os.path.isdir(self.session_dir):
            return

        pasta_base = os.path.dirname(self.path)
        nome_video = os.path.splitext(os.path.basename(self.path))[0]
        momento = datetime.now().strftime("%Y%m%d_%H%M%S")

        self.session_dir = os.path.join(pasta_base, f"{nome_video}_analise_{momento}")
        os.makedirs(self.session_dir, exist_ok=True)

    def abrir(self):
        arquivo, _ = QFileDialog.getOpenFileName(
            self, "Selecionar vídeo", "", "Vídeos (*.mp4 *.avi *.mov *.mkv)"
        )
        if not arquivo:
            return

        if self.cap:
            self.cap.release()

        captura = cv2.VideoCapture(arquivo)
        if not captura.isOpened():
            QMessageBox.critical(self, "Erro", "Não foi possível abrir o vídeo.")
            self.cap = None
            return

        self.cap = captura
        self.path = arquivo
        self.idx = 0
        self.total = int(self.cap.get(cv2.CAP_PROP_FRAME_COUNT))
        self.fps = self.cap.get(cv2.CAP_PROP_FPS) or 0
        self.tab.setRowCount(0)
        self.habilitar_botoes_video(True)
        self.modificado = False

        self.session_dir = None
        self.caminho_projeto = None
        self._criar_pasta_sessao()

        self.mostrar()

    def mostrar(self):
        if not self.cap:
            return

        self.cap.set(cv2.CAP_PROP_POS_FRAMES, self.idx)
        sucesso, frame = self.cap.read()
        if not sucesso:
            return

        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        altura, largura, canais = frame_rgb.shape

        imagem = QImage(
            frame_rgb.data,
            largura,
            altura,
            canais * largura,
            QImage.Format_RGB888
        )

        pixmap = QPixmap.fromImage(imagem).scaled(
            self.lblVideo.width(),
            self.lblVideo.height(),
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation
        )

        self.lblVideo.setPixmap(pixmap)

        tempo = self.idx / self.fps if self.fps else 0
        self.lblInfo.setText(f"Frame: {self.idx} / {self.total - 1}    Tempo: {tempo:.4f} s")

    def prox(self):
        if self.cap and self.idx < self.total - 1:
            self.idx += 1
            self.mostrar()

    def ant(self):
        if self.cap and self.idx > 0:
            self.idx -= 1
            self.mostrar()

    def ir(self):
        if not self.cap:
            return

        valor, confirmado = QInputDialog.getInt(
            self,
            "Ir para...",
            "Frame:",
            value=self.idx,
            min=0,
            max=self.total - 1
        )
        if confirmado:
            self.idx = valor
            self.mostrar()

    def add_linha(self, tipo, comentario):
        tempo = self.idx / self.fps if self.fps else 0
        linha = self.tab.rowCount()
        self.tab.insertRow(linha)

        self.tab.setItem(linha, 0, QTableWidgetItem(tipo))
        self.tab.setItem(linha, 1, QTableWidgetItem(str(self.idx)))
        self.tab.setItem(linha, 2, QTableWidgetItem(f"{tempo:.4f}"))
        self.tab.setItem(linha, 3, QTableWidgetItem(comentario))

        self.modificado = True

    def marcar(self, tipo, titulo, mensagem, obrigatorio=False):
        if not self.cap:
            return

        texto, confirmado = QInputDialog.getText(self, titulo, mensagem)
        if not confirmado:
            return

        if obrigatorio and not texto.strip():
            return

        self.add_linha(tipo, texto)

    def marcar_ini(self):
        self.marcar("inicio", "Início", "Comentário (opcional):")

    def marcar_fim(self):
        self.marcar("fim", "Fim", "Comentário (opcional):")

    def marcar_com(self):
        self.marcar("comentario", "Comentário", "Comentário:", obrigatorio=True)

    def apagar_linha(self):
        linha_atual = self.tab.currentRow()
        if linha_atual < 0:
            return

        resposta = QMessageBox.question(
            self,
            "Confirmar remoção",
            "Deseja realmente apagar esta anotação?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if resposta == QMessageBox.No:
            return

        self.tab.removeRow(linha_atual)
        self.modificado = True

    def salvar_frames(self):
        if not self.path or not self.cap:
            return

        self._criar_pasta_sessao()
        if not self.session_dir:
            QMessageBox.critical(self, "Erro", "Não foi possível determinar a pasta.")
            return

        pasta_frames = os.path.join(self.session_dir, "frames")
        os.makedirs(pasta_frames, exist_ok=True)

        frame_original = self.idx
        self.cap.set(cv2.CAP_PROP_POS_FRAMES, 0)

        fps_local = self.fps or self.cap.get(cv2.CAP_PROP_FPS) or 0
        total_frames = int(self.cap.get(cv2.CAP_PROP_FRAME_COUNT))

        for numero_frame in range(total_frames):
            sucesso, frame = self.cap.read()
            if not sucesso:
                break

            tempo = numero_frame / fps_local if fps_local else 0.0
            nome_arquivo = f"frame_{numero_frame:06d}_t{tempo:010.4f}.jpg"

            cv2.imwrite(
                os.path.join(pasta_frames, nome_arquivo),
                frame,
                [int(cv2.IMWRITE_JPEG_QUALITY), 90]
            )

            if numero_frame % 50 == 0:
                QApplication.processEvents()

        self.cap.set(cv2.CAP_PROP_POS_FRAMES, frame_original)
        self.mostrar()

        QMessageBox.information(
            self,
            "Concluído",
            f"Todos os frames foram salvos em:\n{pasta_frames}"
        )

    def _exportar_xlsx_para(self, caminho_xlsx):
        total_linhas = self.tab.rowCount()

        wb = Workbook()
        ws = wb.active
        ws.title = "Anotacoes"

        colunas = ["Tipo", "Frame", "Tempo(s)", "Comentario"]

        for indice_coluna, texto in enumerate(colunas, start=1):
            ws.cell(row=1, column=indice_coluna, value=texto)

        for linha in range(total_linhas):
            for coluna in range(4):
                item = self.tab.item(linha, coluna)
                conteudo = item.text() if item else ""
                ws.cell(row=linha + 2, column=coluna + 1, value=conteudo)

        wb.save(caminho_xlsx)

    def salvar_xlsx(self):
        if self.tab.rowCount() == 0:
            QMessageBox.information(self, "Sem anotações", "Não há anotações para salvar.")
            return

        self._criar_pasta_sessao()
        if not self.session_dir:
            QMessageBox.critical(self, "Erro", "Não foi possível determinar a pasta.")
            return

        sugestao = os.path.join(self.session_dir, "anotacoes.xlsx")

        arquivo, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar anotações",
            sugestao,
            "Excel (*.xlsx)"
        )
        if not arquivo:
            return

        if not arquivo.lower().endswith(".xlsx"):
            arquivo += ".xlsx"

        try:
            self._exportar_xlsx_para(arquivo)
        except Exception as erro:
            QMessageBox.critical(self, "Erro", f"Erro ao salvar planilha:\n{erro}")
            return

        self.modificado = False
        QMessageBox.information(self, "Concluído", f"Anotações salvas em:\n{arquivo}")

    def salvar_projeto(self):
        if not self.path:
            QMessageBox.information(self, "Sem vídeo", "Abra um vídeo antes de salvar.")
            return

        self._criar_pasta_sessao()
        if not self.session_dir:
            QMessageBox.critical(self, "Erro", "Não foi possível determinar a pasta.")
            return

        sugestao = os.path.join(self.session_dir, "projeto.json")

        arquivo, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar projeto",
            sugestao,
            "Projeto (*.json)"
        )
        if not arquivo:
            return

        if not arquivo.lower().endswith(".json"):
            arquivo += ".json"

        anotacoes = []
        for linha in range(self.tab.rowCount()):
            registro = []
            for coluna in range(4):
                item = self.tab.item(linha, coluna)
                registro.append(item.text() if item else "")
            anotacoes.append(registro)

        dados = {
            "video_path": self.path,
            "frame_atual": self.idx,
            "anotacoes": anotacoes
        }

        try:
            with open(arquivo, "w", encoding="utf-8") as destino:
                json.dump(dados, destino, ensure_ascii=False, indent=2)
        except Exception as erro:
            QMessageBox.critical(self, "Erro", f"Erro ao salvar o projeto:\n{erro}")
            return

        self.caminho_projeto = arquivo
        self.modificado = False

        QMessageBox.information(
            self,
            "Concluído",
            f"Projeto salvo em:\n{arquivo}"
        )

    def _salvar_autosave(self):
        if not self.path or not getattr(self, "modificado", False):
            return

        if not self.caminho_projeto:
            return

        arquivo = self.caminho_projeto

        anotacoes = []
        for linha in range(self.tab.rowCount()):
            registro = []
            for coluna in range(4):
                item = self.tab.item(linha, coluna)
                registro.append(item.text() if item else "")
            anotacoes.append(registro)

        dados = {
            "video_path": self.path,
            "frame_atual": self.idx,
            "anotacoes": anotacoes
        }

        try:
            with open(arquivo, "w", encoding="utf-8") as destino:
                json.dump(dados, destino, ensure_ascii=False, indent=2)
        except Exception:
            pass
        else:
            self.modificado = False

    def abrir_projeto(self):
        arquivo, _ = QFileDialog.getOpenFileName(
            self,
            "Abrir projeto",
            "",
            "Projeto (*.json)"
        )
        if not arquivo:
            return

        try:
            with open(arquivo, "r", encoding="utf-8") as origem:
                dados = json.load(origem)
        except Exception as erro:
            QMessageBox.critical(self, "Erro", f"Não foi possível abrir o projeto:\n{erro}")
            return

        caminho_video = dados.get("video_path")
        if not caminho_video or not os.path.exists(caminho_video):
            QMessageBox.critical(self, "Erro", "O vídeo do projeto não foi encontrado.")
            return

        if self.cap:
            self.cap.release()

        captura = cv2.VideoCapture(caminho_video)
        if not captura.isOpened():
            QMessageBox.critical(self, "Erro", "Erro ao abrir o vídeo do projeto.")
            self.cap = None
            return

        self.cap = captura
        self.path = caminho_video
        self.total = int(self.cap.get(cv2.CAP_PROP_FRAME_COUNT))
        self.fps = self.cap.get(cv2.CAP_PROP_FPS) or 0
        self.idx = dados.get("frame_atual", 0)

        if self.idx < 0:
            self.idx = 0
        if self.idx >= self.total:
            self.idx = max(0, self.total - 1)

        self.tab.setRowCount(0)

        for registro in dados.get("anotacoes", []):
            linha = self.tab.rowCount()
            self.tab.insertRow(linha)

            for coluna in range(4):
                valor = registro[coluna] if coluna < len(registro) else ""
                self.tab.setItem(linha, coluna, QTableWidgetItem(valor))

        self.habilitar_botoes_video(True)
        self.modificado = False
        self.session_dir = os.path.dirname(arquivo)
        self.caminho_projeto = arquivo

        self.mostrar()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyleSheet(qdarktheme.load_stylesheet())
    janela = App()
    janela.show()
    sys.exit(app.exec_())